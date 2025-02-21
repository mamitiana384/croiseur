import pandas as pd
import streamlit as st
import pickle
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import os
import json
import time
from fuzzywuzzy import process
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.filters import AutoFilter
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
# Charger le modèle entraîné et le vectorizer
@st.cache_resource
def load_model():
    with open("model1.pkl", "rb") as f:
        model = pickle.load(f)
    with open("vectorizer.pkl", "rb") as f:
        vectorizer = pickle.load(f)
    return model, vectorizer

model, vectorizer = load_model()

def load_activity_history():
    try:
        if os.path.exists("activity_history.json"):
            if os.stat("activity_history.json").st_size == 0:
                return []
            with open("activity_history.json", "r") as f:
                return json.load(f)
        return []
    except json.JSONDecodeError as e:
        print(f"Erreur de décodage JSON: {e}")
        return []
    except FileNotFoundError:
        return []

def save_to_activity_history(main_filename, ref_filename, columns_crossed, similarity_score):
    activity_history = load_activity_history()
    new_entry = {
        "main_file": main_filename,
        "ref_file": ref_filename,
        "columns_crossed": columns_crossed,
        "similarity_score": int(similarity_score),
        "timestamp": pd.to_datetime("now").strftime('%Y-%m-%d %H:%M:%S')
    }
    activity_history.append(new_entry)
    print(f"activity_history avant sauvegarde: {activity_history}")
    for entry in activity_history:
        print(f"Entry: {entry}")
    with open("activity_history.json", "w") as f:
        json.dump(activity_history, f)

# Afficher l'historique des croisements
def display_activity_history():
    activity_history = load_activity_history()
    if activity_history:
        st.subheader("Historique des activités")
        df_history = pd.DataFrame(activity_history)
        df_history["similarity_score"] = df_history.get("similarity_score", "N/A")
        st.dataframe(df_history)
    else:
        st.write("Aucune activité historique.")

# Charger un fichier Excel et sélectionner un onglet
def load_excel(file):
    xls = pd.ExcelFile(file)
    sheet = st.selectbox(f"Choisissez un onglet pour {file.name}", xls.sheet_names)
    return pd.read_excel(xls, sheet_name=sheet), sheet

# Normalisation des données (nettoyage des valeurs)
def normalize_data(df, column):
    return df[column].astype(str).str.lower().str.strip()

# Trouver la meilleure correspondance entre colonnes avec ML
def find_best_match_ml(main_df, ref_df):
    best_score, best_main_col, best_ref_col = 0, None, None
    
    for main_col in main_df.columns:
        for ref_col in ref_df.columns:
            main_data = normalize_data(main_df, main_col).dropna()
            ref_data = normalize_data(ref_df, ref_col).dropna()

            if main_data.empty or ref_data.empty:
                continue

            main_text = ' '.join(main_data)
            ref_text = ' '.join(ref_data)

            # Transformation avec TF-IDF
            vectors = vectorizer.transform([main_text, ref_text])
            features = vectors.toarray()

            # Prédiction avec le modèle RandomForest
            score = model.predict([features[0] - features[1]])[0] * 100

            if score > best_score:
                best_score, best_main_col, best_ref_col = score, main_col, ref_col

    return best_main_col, best_ref_col, round(best_score, 2)

def cross_check_ml(main_df, ref_df, main_col, ref_col):
    ref_values = set(ref_df[ref_col].astype(str))
    observations, remarks = [], []
    total = len(main_df)
    progress_bar = st.progress(0)

    # Ajouter les colonnes de ref_df à main_df
    for col in ref_df.columns:
        main_df[f"Référence_{col}"] = ""

    for index, (i, value_main) in enumerate(normalize_data(main_df, main_col).items()):
        match, score = process.extractOne(value_main, ref_values) if value_main else (None, 0)

        if score >= 90:
            observations.append("Trouvé")
            remarks.append(f"Correspondance: {match} (Score: {score})")
            # Remplir les colonnes de ref_df
            match_row = ref_df[ref_df[ref_col].astype(str) == match].iloc[0] if match else None
            if match_row is not None:
                for col in ref_df.columns:
                    main_df.loc[i, f"Référence_{col}"] = match_row[col]
        else:
            observations.append("Non trouvé")
            remarks.append("")

        progress_bar.progress(int((index + 1) / total * 100))
        time.sleep(0.01)

    progress_bar.empty()
    main_df["OBS"] = observations
    main_df["REMARQUES"] = remarks
    return main_df

def format_excel(df, filename, main_col, ref_col, similarity_score):
    wb = Workbook()
    ws = wb.active

    # Style pour les en-têtes
    header_font = Font(bold=True, name="Angsana New")
    header_fill_main = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
    header_fill_obs = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    header_fill_ref = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    alignment = Alignment(horizontal="center")

    # Style pour le corps du tableau
    body_font = Font(name="Angsana New")
    border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"), top=Side(border_style="thin"), bottom=Side(border_style="thin"))

    # Écriture des en-têtes et application des styles
    for i, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=i, value=col)
        cell.font = header_font
        cell.alignment = alignment
        if col in ["OBS", "REMARQUES"]:
            cell.fill = header_fill_obs
        elif col.startswith("Référence_"):
            cell.fill = header_fill_ref
        else:
            cell.fill = header_fill_main

    # Écriture des données et application des styles
    for i, row in df.iterrows():
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=i + 2, column=j, value=str(val))
            cell.font = body_font
            cell.border = border
            cell.alignment = alignment
            if df.columns[j - 1] == "OBS":
                if val == "Trouvé":
                    cell.fill = PatternFill(start_color="00FF00", fill_type="solid")
                elif val == "Non trouvé":
                    cell.fill = PatternFill(start_color="FF0000", fill_type="solid")
            # Formatage conditionnel pour les différences
            if df.columns[j - 1].startswith("Référence_"):
                original_col_name = df.columns[j - 1][10:]
                if original_col_name in df.columns:  # Vérification de l'existence de la colonne
                    if row[original_col_name] != val:
                        cell.fill = PatternFill(start_color="FFFF00", fill_type="solid")

    # Ajustement automatique des largeurs de colonnes
    for col in ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Création de l'onglet de résumé
    summary_ws = wb.create_sheet("Résumé")

    # Mise en forme des en-têtes du résumé
    for col_num, header_text in enumerate(["Paramètre", "Valeur"], start=1):
        cell = summary_ws.cell(row=1, column=col_num, value=header_text)
        cell.font = header_font
        cell.alignment = alignment
        cell.fill = header_fill_main  # Utiliser la même couleur que les en-têtes principaux

    # Mise en forme des données du résumé
    summary_data = {
        "Colonne principale": main_col,
        "Colonne de référence": ref_col,
        "Score de similarité": similarity_score,
        "Correspondances trouvées": len(df[df["OBS"] == "Trouvé"]),
        "Correspondances non trouvées": len(df[df["OBS"] == "Non trouvé"]),
    }

    for row_num, (param, value) in enumerate(summary_data.items(), start=2):
        summary_ws.cell(row=row_num, column=1, value=param).font = body_font
        summary_ws.cell(row=row_num, column=2, value=str(value)).font = body_font
        summary_ws.cell(row=row_num, column=1).alignment = Alignment(horizontal="left")
        summary_ws.cell(row=row_num, column=2).alignment = alignment
        summary_ws.cell(row=row_num, column=1).border = border
        summary_ws.cell(row=row_num, column=2).border = border

    # Ajustement automatique des largeurs de colonnes du résumé
    for col in summary_ws.columns:
        max_length = 0
        column = get_column_letter(col[0].column)
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except TypeError:
                pass
        adjusted_width = (max_length + 2)
        summary_ws.column_dimensions[column].width = adjusted_width

    # Ajout des filtres
    ws.auto_filter.ref = ws.dimensions

    wb.save(filename)

# Interface principale Streamlit
def main():
    st.title("Croisement de fichiers Excel avec Machine Learning 🔥")
    
    tabs = st.sidebar.radio("Sélectionner un onglet", ["Croisement de fichiers", "Historique des activités"])
    
    if tabs == "Croisement de fichiers":
        main_file = st.sidebar.file_uploader("Chargez le fichier principal", type=["xlsx"])
        ref_file = st.sidebar.file_uploader("Chargez le fichier de référence", type=["xlsx"])
        
        if main_file and ref_file:
            main_df, main_sheet = load_excel(main_file)
            ref_df, ref_sheet = load_excel(ref_file)
            
            best_main_col, best_ref_col, similarity_score = find_best_match_ml(main_df, ref_df)
            st.write(f"**Colonne suggérée:** {best_main_col} vs {best_ref_col} (Score: {similarity_score}%)")
            
            col1, col2 = st.columns(2)
            with col1:
                main_col = st.selectbox("Colonne principale", main_df.columns, index=main_df.columns.get_loc(best_main_col))
            with col2:
                ref_col = st.selectbox("Colonne de référence", ref_df.columns, index=ref_df.columns.get_loc(best_ref_col))
            
            if st.button("Lancer le croisement"):
                with st.spinner("Croisement en cours..."):
                    result_df = cross_check_ml(main_df, ref_df, main_col, ref_col)
                st.success("Croisement terminé !")
                st.dataframe(result_df)

                output_filename = f"croisement_{main_file.name.split('.')[0]}_{ref_file.name.split('.')[0]}.xlsx"
                format_excel(result_df, output_filename, main_col, ref_col, similarity_score) # Ajout des arguments ici
                with open(output_filename, "rb") as f:
                    st.download_button("Télécharger le fichier Excel", f, file_name=output_filename, mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

                save_to_activity_history(main_file.name, ref_file.name, f"{main_col} vs {ref_col}", similarity_score)
    
    elif tabs == "Historique des activités":
        display_activity_history()

if __name__ == "__main__":
    main()
